"""Compilazione componenti VLD – programma di inserimento dati su Excel.

Consente di selezionare un file Excel, cercare una matricola nella
colonna A del foglio 1, e compilare i campi dei componenti associati
(diodo, tiristore, varistore, scheda controllo).
"""

import sys
import os

from PyQt5 import QtWidgets, QtCore, QtGui
import openpyxl


# ---------------------------------------------------------------------------
# Mappatura colonne Excel (1-indexed, come openpyxl)
# ---------------------------------------------------------------------------
#   A = 1  → Matricola (ricerca)
#   B,C,D  → Diodo       (tipo, produttore, n. seriale)
#   E,F,G  → Tiristore   (tipo, produttore, n. seriale)
#   H,I,J  → Varistore   (tipo, produttore, n. seriale)
#   K,L,M  → Scheda ctrl (tipo, produttore, n. seriale)

COMPONENTS = [
    ("Diodo",            2),   # colonna B
    ("Tiristore",        5),   # colonna E
    ("Varistore",        8),   # colonna H
    ("Scheda Controllo", 11),  # colonna K
]
MATRICOLA_COL = 1  # colonna A

# Serial number split configuration per component.
# fmt      : how the two positional parts are joined when written to Excel
# sep      : separator used to split the raw value when reading
# date_idx : which positional part (0 or 1) holds the date/lotto
SERIAL_FORMATS = {
    "Diodo":            {"fmt": "{0} - {1}", "sep": " - ", "date_idx": 0},
    "Tiristore":        {"fmt": "{0} - {1}", "sep": " - ", "date_idx": 0},
    "Varistore":        {"fmt": "{0} {1}",   "sep": " ",   "date_idx": 1},
    "Scheda Controllo": {"fmt": "{0} - {1}", "sep": " - ", "date_idx": 1},
}

# Zero-padding width for the serial number part (0 = no padding)
SERIAL_PAD = {
    "Diodo": 3,            # "1"  → "001"
    "Tiristore": 3,        # "44" → "044"
    "Varistore": 0,        # no padding (prefix handled separately)
    "Scheda Controllo": 4, # "1"  → "0001"
}

# Varistore serial is split into a persistent prefix + number
# e.g. "D9514866" → prefix "D95" + number "14866"
VARISTORE_PREFIX_DEFAULT = "D95"
VARISTORE_PREFIX_LEN = 3  # characters to split off when parsing


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Compilazione Componenti VLD")
        self.setMinimumWidth(1100)
        self._current_row: int | None = None
        self._setup_ui()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _setup_ui(self):
        self.setStyleSheet("""
            * { font-size: 18px; }
            QMainWindow { background: #f0f2f5; }
            QGroupBox {
                font-weight: bold; font-size: 20px;
                border: 1px solid #bbb;
                border-radius: 4px;
                margin-top: 14px;
                padding-top: 18px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 4px;
            }
            QGroupBox#comp_group {
                background: #fafafa;
            }
            QLineEdit { font-size: 18px; padding: 4px; }
            QLineEdit[objectName^="serial_"] {
                background-color: #FFFDE7;
                border: 2px solid #FBC02D;
                border-radius: 3px;
            }
            QLineEdit[objectName^="serial_"]:focus {
                border: 2px solid #F57F17;
                background-color: #FFF9C4;
            }
            QLabel { font-size: 18px; }
            QPushButton { font-size: 18px; padding: 6px 14px; }
            QPushButton#save_btn {
                background-color: #1976D2; color: white;
                font-weight: bold; padding: 10px 24px;
                border: none; border-radius: 4px;
                font-size: 20px;
            }
            QPushButton#save_btn:hover { background-color: #1565C0; }
            QPushButton#save_btn:disabled { background-color: #aaa; }
            QPushButton#save_next_btn {
                background-color: #4CAF50; color: white;
                font-weight: bold; padding: 10px 24px;
                border: none; border-radius: 4px;
                font-size: 20px;
            }
            QPushButton#save_next_btn:hover { background-color: #43A047; }
            QPushButton#save_next_btn:disabled { background-color: #aaa; }
        """)

        central = QtWidgets.QWidget()
        main_layout = QtWidgets.QVBoxLayout()

        # --- File Excel ---
        file_group = QtWidgets.QGroupBox("File Excel")
        file_lay = QtWidgets.QHBoxLayout()
        self.excel_path_edit = QtWidgets.QLineEdit()
        self.excel_path_edit.setPlaceholderText("Seleziona il file Excel…")
        browse_btn = QtWidgets.QPushButton("Sfoglia…")
        browse_btn.clicked.connect(self._browse_excel)
        file_lay.addWidget(self.excel_path_edit, 1)
        file_lay.addWidget(browse_btn)
        file_group.setLayout(file_lay)
        main_layout.addWidget(file_group)

        # --- Matricola ---
        matr_group = QtWidgets.QGroupBox("Matricola")
        matr_lay = QtWidgets.QHBoxLayout()
        self.matricola_edit = QtWidgets.QLineEdit()
        self.matricola_edit.setPlaceholderText("Inserisci la matricola…")
        self.matricola_edit.textChanged.connect(self._search_matricola)
        minus_btn = QtWidgets.QPushButton("\u2212")  # −
        minus_btn.setFixedWidth(44)
        minus_btn.clicked.connect(self._matricola_decrement)
        plus_btn = QtWidgets.QPushButton("+")
        plus_btn.setFixedWidth(44)
        plus_btn.clicked.connect(self._matricola_increment)
        matr_lay.addWidget(self.matricola_edit, 1)
        matr_lay.addWidget(minus_btn)
        matr_lay.addWidget(plus_btn)
        matr_group.setLayout(matr_lay)
        main_layout.addWidget(matr_group)

        # --- Status label ---
        self.status_label = QtWidgets.QLabel("")
        self.status_label.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addWidget(self.status_label)

        # --- Component fields (grid: rows = components, cols = fields) ---
        comp_group = QtWidgets.QGroupBox("Componenti")
        comp_group.setObjectName("comp_group")
        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        # Column headers
        header_colors = ["", "", "", "#F57F17", "#E65100"]
        for col, header in enumerate(["", "Tipo", "Produttore", "Lotto", "N. Seriale"]):
            lbl = QtWidgets.QLabel(header)
            style = "font-weight: bold;"
            if header_colors[col]:
                style += f" color: {header_colors[col]};"
            lbl.setStyleSheet(style)
            if col == 0:
                lbl.setFixedWidth(140)
            grid.addWidget(lbl, 0, col)
        # Row label colors per component
        ROW_COLORS = ["#1565C0", "#2E7D32", "#BF360C", "#4A148C"]
        # One row per component
        self._comp_edits: dict[str, dict[str, QtWidgets.QLineEdit]] = {}
        for row_idx, (comp_name, _) in enumerate(COMPONENTS, start=1):
            row_label = QtWidgets.QLabel(comp_name)
            row_label.setStyleSheet(
                f"font-weight: bold; color: {ROW_COLORS[row_idx - 1]};")
            grid.addWidget(row_label, row_idx, 0)
            edits: dict[str, QtWidgets.QLineEdit] = {}
            for col_idx, key in enumerate(("tipo", "prod", "lotto"), start=1):
                edit = QtWidgets.QLineEdit()
                edit.setEnabled(False)
                grid.addWidget(edit, row_idx, col_idx)
                edits[key] = edit
            # Serial column — varistore gets prefix + number side-by-side
            if comp_name == "Varistore":
                container = QtWidgets.QWidget()
                hlay = QtWidgets.QHBoxLayout(container)
                hlay.setContentsMargins(0, 0, 0, 0)
                hlay.setSpacing(4)
                prefix_edit = QtWidgets.QLineEdit(VARISTORE_PREFIX_DEFAULT)
                prefix_edit.setFixedWidth(60)
                prefix_edit.setEnabled(False)
                prefix_edit.setObjectName("prefix_Varistore")
                hlay.addWidget(prefix_edit)
                serial_edit = QtWidgets.QLineEdit()
                serial_edit.setEnabled(False)
                serial_edit.setObjectName(f"serial_{comp_name}")
                hlay.addWidget(serial_edit)
                grid.addWidget(container, row_idx, 4)
                edits["prefix"] = prefix_edit
                edits["serial"] = serial_edit
            else:
                serial_edit = QtWidgets.QLineEdit()
                serial_edit.setEnabled(False)
                serial_edit.setObjectName(f"serial_{comp_name}")
                grid.addWidget(serial_edit, row_idx, 4)
                edits["serial"] = serial_edit
            self._comp_edits[comp_name] = edits
        grid.setColumnStretch(0, 0)
        grid.setColumnStretch(1, 3)
        grid.setColumnStretch(2, 2)
        grid.setColumnStretch(3, 2)
        grid.setColumnStretch(4, 3)
        comp_group.setLayout(grid)
        main_layout.addWidget(comp_group)

        # --- Save buttons ---
        btn_lay = QtWidgets.QHBoxLayout()
        self.save_btn = QtWidgets.QPushButton("Salva su Excel")
        self.save_btn.setObjectName("save_btn")
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self._save)
        btn_lay.addWidget(self.save_btn)

        self.save_next_btn = QtWidgets.QPushButton(
            "Salva e passa alla matricola successiva")
        self.save_next_btn.setObjectName("save_next_btn")
        self.save_next_btn.setEnabled(False)
        self.save_next_btn.clicked.connect(self._save_and_next)
        btn_lay.addWidget(self.save_next_btn)
        main_layout.addLayout(btn_lay)

        central.setLayout(main_layout)
        self.setCentralWidget(central)

        # --- Tab order: serial fields → Save & Next ---
        serial_widgets = [self._comp_edits[name]["serial"]
                          for name, _ in COMPONENTS]
        for i in range(len(serial_widgets) - 1):
            self.setTabOrder(serial_widgets[i], serial_widgets[i + 1])
        self.setTabOrder(serial_widgets[-1], self.save_next_btn)

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------
    def _browse_excel(self):
        default_dir = r"Y:\Projects\Produzione\9453 - VLD RFI"
        if not os.path.isdir(default_dir):
            default_dir = "C:\\"
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Seleziona file Excel", default_dir,
            "File Excel (*.xlsx *.xls);;Tutti i file (*)",
        )
        if path:
            self.excel_path_edit.setText(path)
            # Reset state when file changes
            self._current_row = None
            self._set_fields_enabled(False)
            self._clear_fields()
            self.status_label.setText("")

    def _search_matricola(self):
        path = self.excel_path_edit.text().strip()
        matricola = self.matricola_edit.text().strip()

        if not path or not matricola:
            self._current_row = None
            self._set_fields_enabled(False)
            self._clear_fields()
            self.status_label.setText("")
            return

        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Errore", f"Impossibile aprire il file:\n{e}")
            return

        ws = wb.worksheets[0]
        target_row = None
        for row in ws.iter_rows(min_col=MATRICOLA_COL,
                                max_col=MATRICOLA_COL):
            cell = row[0]
            if str(cell.value).strip() == matricola:
                target_row = cell.row
                break

        if target_row is None:
            wb.close()
            self._current_row = None
            self._set_fields_enabled(False)
            self._clear_fields()
            self.status_label.setText(
                f"<span style='color:red;'>Matricola «{matricola}» "
                f"non trovata.</span>")
            return

        # Found — load existing data into fields
        self._current_row = target_row
        for comp_name, col_start in COMPONENTS:
            ce = self._comp_edits[comp_name]
            tipo_val = ws.cell(row=target_row, column=col_start).value
            prod_val = ws.cell(row=target_row, column=col_start + 1).value
            # Only overwrite tipo/prod if Excel cell has data; otherwise
            # keep the previous value (persistent across matricole).
            if tipo_val is not None:
                ce["tipo"].setText(str(tipo_val))
            if prod_val is not None:
                ce["prod"].setText(str(prod_val))
            # Parse serial into lotto (date) and serial parts
            raw = ws.cell(row=target_row, column=col_start + 2).value
            sf = SERIAL_FORMATS[comp_name]
            if raw:
                parts = str(raw).split(sf["sep"], 1)
                di = sf["date_idx"]
                # Only overwrite lotto if a date part was actually parsed
                if len(parts) > di and parts[di].strip():
                    ce["lotto"].setText(parts[di].strip())
                serial_part = parts[1 - di].strip() if len(parts) > 1 else ""
                # Varistore: split prefix from serial number
                if comp_name == "Varistore" and serial_part:
                    ce["prefix"].setText(
                        serial_part[:VARISTORE_PREFIX_LEN] or VARISTORE_PREFIX_DEFAULT)
                    ce["serial"].setText(serial_part[VARISTORE_PREFIX_LEN:])
                else:
                    ce["serial"].setText(serial_part)
            else:
                # Lotto, prefix, tipo, prod persist from previous matricola
                ce["serial"].setText("")

        wb.close()
        self._set_fields_enabled(True)
        self.status_label.setText(
            f"<span style='color:green;'>Matricola «{matricola}» trovata "
            f"alla riga {target_row}.</span>")
        # Focus on Diodo serial for quick data entry
        self._comp_edits[COMPONENTS[0][0]]["serial"].setFocus()

    def _save(self) -> bool:
        """Write current fields to Excel. Returns *True* on success."""
        path = self.excel_path_edit.text().strip()
        if not path or self._current_row is None:
            return False

        try:
            wb = openpyxl.load_workbook(path)
        except Exception:
            QtWidgets.QMessageBox.critical(
                self, "Errore",
                "Impossibile aprire il file Excel.\n\n"
                "Il file potrebbe essere già aperto su questo PC "
                "o sul PC di un collega.\n"
                "Chiudere il file e riprovare.")
            return False

        ws = wb.worksheets[0]
        row = self._current_row

        # --- Duplicate serial check (per component) ---
        duplicates: list[str] = []
        for comp_name, col_start in COMPONENTS:
            ce = self._comp_edits[comp_name]
            sf = SERIAL_FORMATS[comp_name]
            lotto = ce["lotto"].text().strip()
            serial = ce["serial"].text().strip()
            prod = ce["prod"].text().strip()
            if not serial or not prod:
                continue
            # Apply zero-padding
            pad = SERIAL_PAD.get(comp_name, 0)
            if pad:
                try:
                    serial = str(int(serial)).zfill(pad)
                except ValueError:
                    pass
            # Varistore: prepend prefix to serial
            if comp_name == "Varistore":
                prefix_edit = ce.get("prefix")
                prefix_text = prefix_edit.text().strip() if prefix_edit else ""
                serial = prefix_text + serial
            # Build the combined value that will be written
            vals = ["", ""]
            vals[sf["date_idx"]] = lotto
            vals[1 - sf["date_idx"]] = serial
            new_combined = sf["fmt"].format(vals[0], vals[1])
            # Scan the column for existing values with same produttore
            serial_col = col_start + 2
            prod_col = col_start + 1
            for scan_row in ws.iter_rows(min_row=2, min_col=1, max_col=max(serial_col, prod_col)):
                r = scan_row[0].row
                if r == row:
                    continue  # skip own row
                existing_prod = scan_row[prod_col - 1].value
                if not existing_prod or str(existing_prod).strip() != prod:
                    continue
                existing_serial = scan_row[serial_col - 1].value
                if existing_serial and str(existing_serial).strip() == new_combined:
                    matr_cell = scan_row[0].value
                    duplicates.append(
                        f"  • {comp_name}: \"{new_combined}\" già presente "
                        f"alla matricola {matr_cell} (riga {r})")

        if duplicates:
            msg = ("Attenzione: numeri seriali duplicati!\n\n"
                   + "\n".join(duplicates)
                   + "\n\nVuoi salvare comunque?")
            reply = QtWidgets.QMessageBox.warning(
                self, "Seriale duplicato", msg,
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No)
            if reply != QtWidgets.QMessageBox.Yes:
                wb.close()
                return False

        # --- Write fields ---

        for comp_name, col_start in COMPONENTS:
            ce = self._comp_edits[comp_name]
            ws.cell(row=row, column=col_start,
                    value=ce["tipo"].text().strip() or None)
            ws.cell(row=row, column=col_start + 1,
                    value=ce["prod"].text().strip() or None)
            # Join lotto + serial into the configured format
            sf = SERIAL_FORMATS[comp_name]
            lotto = ce["lotto"].text().strip()
            serial = ce["serial"].text().strip()
            # Apply zero-padding
            pad = SERIAL_PAD.get(comp_name, 0)
            if pad and serial:
                try:
                    serial = str(int(serial)).zfill(pad)
                except ValueError:
                    pass
            # Varistore: prepend prefix
            if comp_name == "Varistore":
                prefix_edit = ce.get("prefix")
                serial = (prefix_edit.text().strip() if prefix_edit else "") + serial
            if lotto or serial:
                vals = ["", ""]
                vals[sf["date_idx"]] = lotto
                vals[1 - sf["date_idx"]] = serial
                combined = sf["fmt"].format(vals[0], vals[1])
            else:
                combined = None
            ws.cell(row=row, column=col_start + 2, value=combined)

        try:
            wb.save(path)
            wb.close()
        except Exception:
            wb.close()
            QtWidgets.QMessageBox.critical(
                self, "Errore",
                "Impossibile salvare il file Excel.\n\n"
                "Il file potrebbe essere già aperto su questo PC "
                "o sul PC di un collega.\n"
                "Chiudere il file e riprovare.")
            return False

        matricola = self.matricola_edit.text().strip()
        QtWidgets.QMessageBox.information(
            self, "Salvato",
            f"Dati per la matricola «{matricola}» salvati correttamente.")
        return True

    def _save_and_next(self):
        """Save current data, then move to the next matricola."""
        if self._save():
            self._matricola_increment()

    def _matricola_increment(self):
        text = self.matricola_edit.text().strip()
        try:
            value = int(text)
        except (ValueError, TypeError):
            value = 0
        self.matricola_edit.setText(str(value + 1))

    def _matricola_decrement(self):
        text = self.matricola_edit.text().strip()
        try:
            value = int(text)
        except (ValueError, TypeError):
            value = 1
        self.matricola_edit.setText(str(max(0, value - 1)))

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _set_fields_enabled(self, enabled: bool):
        for ce in self._comp_edits.values():
            for edit in ce.values():
                edit.setEnabled(enabled)
        self.save_btn.setEnabled(enabled)
        self.save_next_btn.setEnabled(enabled)

    def _clear_fields(self):
        """Clear only what changes per matricola.

        Tipo, Produttore, Lotto and Varistore prefix persist because they
        are typically constant across many consecutive entries.
        Only the serial number is cleared.
        """
        for ce in self._comp_edits.values():
            ce["serial"].clear()


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
