"""Excel file handling for automatic test data recording.

Uses openpyxl to read/write .xlsx files. Each test routine writes
voltage readings and pass/fail markers to the row whose column-A
value matches the operator-supplied matricola.
"""

import openpyxl


class ExcelHandler:
    """Read/write test results to an Excel workbook."""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self.ws_lookup = self.wb.worksheets[0]  # primo foglio — matricole
        self.ws_data = self.wb.worksheets[3]     # quarto foglio — dati

    # ------------------------------------------------------------------
    # Lookup
    # ------------------------------------------------------------------

    @staticmethod
    def _normalise(value: str) -> str:
        """Strip whitespace and leading zeros so '007' matches '7'."""
        s = str(value).strip().lstrip('0')
        return s if s else '0'          # keep at least one '0'

    def find_row_by_matricola(self, matricola: str):
        """Return the 1-based row index whose column-A value matches
        *matricola* (ignoring leading zeros), or ``None`` if not found.
        """
        target = self._normalise(matricola)
        for row in self.ws_lookup.iter_rows(min_col=1, max_col=1, values_only=False):
            cell = row[0]
            if cell.value is not None and self._normalise(cell.value) == target:
                return cell.row
        return None

    # ------------------------------------------------------------------
    # Low-level helpers
    # ------------------------------------------------------------------

    def write_cell(self, row: int, col: int, value):
        """Write *value* to the cell at (*row*, *col*) on the data sheet."""
        self.ws_data.cell(row=row, column=col, value=value)

    def save(self):
        """Persist changes to disk.

        Raises a clear ``PermissionError`` if the file is locked
        (e.g. open in Excel).
        """
        try:
            self.wb.save(self.filepath)
        except PermissionError:
            raise PermissionError(
                f"Impossibile salvare '{self.filepath}'.\n"
                "Il file potrebbe essere aperto in Excel.\n"
                "Chiudere il file e riprovare."
            )

    def close(self):
        """Close the workbook (does **not** save automatically)."""
        self.wb.close()

    # ------------------------------------------------------------------
    # Per-routine writers
    # ------------------------------------------------------------------

    def write_innesco_results(self, row: int, diode_drop: float,
                              trigger_voltage: float) -> list:
        """Write *Innesco Tiristore* results.

        * Column H (8)  – caduta di tensione del tiristore
          If < 5 → Column J (10) = "POS."
        * Column K (11) – tensione di innesco del tiristore
          If 90–120 → Column M (13) = "OK", Column N (14) = "POS."

        Returns a list of error strings (empty when everything is OK).
        """
        errors = []

        # --- caduta di tensione ---
        if diode_drop < 5:
            self.write_cell(row, 8, round(diode_drop, 2))   # H
            self.write_cell(row, 10, "POS.")                 # J
        else:
            errors.append(
                f"Caduta di tensione tiristore ({diode_drop:.2f} V) "
                f"non inferiore a 5 V"
            )

        # --- tensione di innesco ---
        if 90 <= trigger_voltage <= 120:
            self.write_cell(row, 11, int(round(trigger_voltage)))  # K — senza decimali
            self.write_cell(row, 13, "OK")                         # M
            self.write_cell(row, 14, "POS.")                       # N
        else:
            errors.append(
                f"Tensione innesco ({trigger_voltage:.2f} V) "
                f"fuori range 90–120 V"
            )

        self.save()
        return errors

    def write_at_al_results(self, row: int, voltage: float) -> list:
        """Write *Anomalia Tiristore e Limiti (AT + AL)* results.

        * Column O (15) – tensione
          If 90–120 → Column Q (17) = "OK", Column R (18) = "OK",
                       Column S (19) = "POS."

        Returns a list of error strings.
        """
        errors = []

        if 90 <= voltage <= 120:
            self.write_cell(row, 15, round(voltage, 2))  # O
            self.write_cell(row, 17, "OK")               # Q
            self.write_cell(row, 18, "OK")               # R
            self.write_cell(row, 19, "POS.")             # S
        else:
            errors.append(
                f"Tensione AT+AL ({voltage:.2f} V) fuori range 90–120 V"
            )

        self.save()
        return errors

    def write_ad_results(self, row: int, voltage: float) -> list:
        """Write *Anomalia Diodo (AD)* results.

        * Column T (20) – tensione
          If 90–120 → Column V (22) = "POS."

        Returns a list of error strings.
        """
        errors = []

        if 90 <= voltage <= 120:
            self.write_cell(row, 20, round(voltage, 2))  # T
            self.write_cell(row, 22, "POS.")             # V
        else:
            errors.append(
                f"Tensione AD ({voltage:.2f} V) fuori range 90–120 V"
            )

        self.save()
        return errors

    def write_100v_results(self, row: int, current: float) -> list:
        """Write *Prova 100 V* results.

        If current > 1.9 A:
        * Column B (2) = "OK"
        * Column C (3) = "POS."
        * Column D (4) = "OK"
        * Column E (5) = "POS."

        Returns a list of error strings.
        """
        errors = []

        if current > 1.9:
            self.write_cell(row, 2, "OK")    # B
            self.write_cell(row, 3, "POS.")  # C
            self.write_cell(row, 4, "OK")    # D
            self.write_cell(row, 5, "POS.")  # E
        else:
            errors.append(
                f"Corrente misurata ({current:.3f} A) non superiore a 1.9 A"
            )

        self.save()
        return errors

    def write_500v_results(self, row: int, current: float) -> list:
        """Write *Prova 500 V* results.

        If current <= 0.005 A:
        * Column F (6) = "OK"
        * Column G (7) = "POS."

        Returns a list of error strings (includes current value if > 0.005).
        """
        errors = []

        if current <= 0.005:
            self.write_cell(row, 6, "OK")    # F
            self.write_cell(row, 7, "POS.")  # G
        else:
            errors.append(
                f"Corrente misurata: {current:.3f} A — Problema al Tiristore"
            )

        self.save()
        return errors

    def get_test_status(self, row: int) -> dict:
        """Return {test_name: bool} — True if that test has been recorded.

        Uses the primary result cell for each test as presence indicator:
          100V    → col B  (2)
          500V    → col F  (6)
          Innesco → col H  (8)
          AT+AL   → col O  (15)
          AD      → col T  (20)
        """
        def filled(col):
            return self.ws_data.cell(row, col).value is not None
        return {
            '100V':    filled(2),
            '500V':    filled(6),
            'Innesco': filled(8),
            'AT+AL':   filled(15),
            'AD':      filled(20),
        }
